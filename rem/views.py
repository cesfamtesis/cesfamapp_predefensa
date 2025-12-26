from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.core.paginator import Paginator
from django.db import transaction
from django.utils import timezone
from datetime import date
from collections import Counter, defaultdict
import os
from decimal import Decimal
from django.db.models import Count

from io import BytesIO
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors


from django.contrib.auth.decorators import login_required
from rem.decorators import admin_required

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side

from .models import DimPeriodo, ArchivoREM, RegistroREM, AuditLog
from .etl import procesar_archivo_con_mapeo
from rem.auditoria import registrar_auditoria

from openpyxl import load_workbook
from django.contrib import messages



# ============================================================
# CARGA DE ESTRUCTURAS REM (JSON / PY) PARA HEADERS FIJOS
# ============================================================
# Si existe rem_structures.py, se usa para definir columnas por REM/Sección.
# Si no existe, el sistema funciona igual pero "descubre" columnas desde datos.
try:
    from .rem_structures import REM_STRUCTURES
except ImportError:
    REM_STRUCTURES = {}


# ============================================================
# METADATOS: RANGOS / HEADERS ESPECIALES (REM A01 - SECCIÓN A)
# ============================================================
# RANGOS etarios usados para reportes, export Excel y export PDF.
RANGOS_A01_A = [
    ("rango_etario_menos_de_4_anos", "Menos de 4 años"),
    ("rango_etario_5_9_anos", "5 - 9 años"),
    ("rango_etario_10_14_anos", "10 - 14 años"),
    ("rango_etario_15_19_anos", "15 - 19 años"),
    ("rango_etario_20_24_anos", "20 - 24 años"),
    ("rango_etario_25_29_anos", "25 - 29 años"),
    ("rango_etario_30_34_anos", "30 - 34 años"),
    ("rango_etario_35_39_anos", "35 - 39 años"),
    ("rango_etario_40_44_anos", "40 - 44 años"),
    ("rango_etario_45_49_anos", "45 - 49 años"),
    ("rango_etario_50_54_anos", "50 - 54 años"),
    ("rango_etario_55_59_anos", "55 - 59 años"),
    ("rango_etario_60_64_anos", "60 - 64 años"),
    ("rango_etario_65_69_anos", "65 - 69 años"),
    ("rango_etario_70_74_anos", "70 - 74 años"),
    ("rango_etario_75_79_anos", "75 - 79 años"),
    ("rango_etario_80_y_mas_anos", "80 y más años"),
]

# Nombres legibles para columnas especiales de A01 / Sección A
# (Los rangos etarios se resuelven desde RANGOS_A01_A_DICT)
HEADERS_A01_A = {
    "tipo_de_control": "Tipo de control",
    "profesional": "Profesional",
    "total": "TOTAL",

    "sexo_hombres": "Hombres",
    "sexo_mujeres": "Mujeres",

    "control_con_pareja_familiar_u_otro": "Control con pareja, familiar u otro",
    "control_de_diada_con_presencia_del_padre": "Control de diada con presencia del padre",
    "espacios_amigables_adolescentes": "Espacios Amigables/Adolescentes",
    "nna_sename": "Niños, Niñas, Adolescentes y Jóvenes SENAME",
    "nna_mejor_ninez": "Niños, Niñas, Adolescentes y Jóvenes Mejor Niñez",
    "pueblos_originarios": "Pueblos Originarios",
    "migrantes": "Migrantes",
    "personas_con_discapacidad": "Personas con discapacidad",

    "identificacion_de_genero_trans_masculino": "Trans masculino",
    "identificacion_de_genero_trans_femenina": "Trans femenina",
    "identificacion_de_genero_no_binarie": "No binarie",

    "adolescente_acude_a_control_mac_con_pareja": "Adolescente acude a control MAC con pareja",
}

# Dict rápido (key -> label) para resolver nombres de rangos etarios
RANGOS_A01_A_DICT = {k: v for k, v in RANGOS_A01_A}

# Nombres bonitos para mostrar en combos / títulos de REM
REM_TITULOS_HOJA = {
    "A01": "REM-A01. CONTROLES DE SALUD",
    "A02": "REM-A02. EXAMEN DE MEDICINA PREVENTIVA EN MAYORES DE 15 AÑOS",
    "A03": "REM-A03. APLICACIÓN Y RESULTADOS DE ESCALAS DE EVALUACIÓN",
    "A04": "REM-A04. CONSULTAS Y OTRAS ATENCIONES",
    "A05": "REM-A05. INGRESOS Y EGRESOS POR CONDICIÓN Y PROBLEMAS DE SALUD",
    "A06": "REM-A06. PROGRAMA DE SALUD MENTAL ATENCIÓN PRIMARIA Y ESPECIALIDADES",
    "A07": "REM-A07. ATENCIÓN DE ESPECIALIDADES",
    "A08": "REM-A08. ATENCIÓN DE URGENCIA",
    "A09": "REM-A09. ATENCIÓN DE SALUD ODONTOLÓGICA",
    "A11": "REM-A11. EXÁMENES DE PESQUISA DE ENFERMEDADES TRASMISIBLES ",
    "A11A": "REM-A11A. TRANSMISIÓN VERTICAL MATERNO INFANTIL",
    "A19A": "REM-A19A. ACTIVIDADES DE PROMOCIÓN Y PREVENCIÓN DE LA SALUD",
    "A19B": "REM-A19B. ACTIVIDADES DE PARTICIPACIÓN SOCIAL",
    "A23": "REM-A23. SALAS: IRA, ERA Y MIXTAS EN APS",
    "A26": "REM-A26. ACTIVIDADES EN DOMICILIO Y OTROS ESPACIOS",
    "A27": "REM-A27. EDUCACIÓN PARA LA SALUD",
    "A28": "REM-A28. REHABILITACIÓN INTEGRAL",
    "A29": "REM-A29. PROGRAMA DE IMÁGENES DIAGNÓSTICAS Y/O RESOLUTIVIDAD EN ATENCIÓN PRIMARIA",
    "A30": "REM-A30. ATENCIONES POR TELEMEDICINA EN LA RED ASISTENCIAL",
    "A32": "REM-A32. ACTIVIDADES DE ATENCIÓN DE SALUD REMOTA",
    "A33": "REM-A33. CUIDADOS PALIATIVOS NIVEL APS Y HOSPITALARIO",
    # agrega más si usa más hojas
}


# ========================
# Home
# ========================
@login_required
def home(request):
    """
    Página principal del sistema.
    Normalmente aquí se muestran accesos rápidos a módulos (subir REM, periodos, reportes, auditoría, etc.).
    """
    return render(request, "home.html")


# ========================
# MÓDULO 3: GESTIÓN DE PERÍODOS (CRUD)
# ========================
@login_required
def lista_periodos(request):
    """
    Lista los períodos (DimPeriodo) separando:
    - Activos   : sin la marca [INACTIVO] en su descripción
    - Inactivos : con la marca [INACTIVO] en su descripción

    Nota: tu sistema usa "borrado lógico" para períodos, por eso se marca [INACTIVO]
    en vez de eliminar registros.
    """
    periodos_activos = (
        DimPeriodo.objects
        .exclude(descripcion__icontains='[INACTIVO]')
        .order_by('-anio', '-mes')
    )

    periodos_inactivos = (
        DimPeriodo.objects
        .filter(descripcion__icontains='[INACTIVO]')
        .order_by('-anio', '-mes')
    )

    return render(request, "lista_periodos.html", {
        "periodos_activos": periodos_activos,
        "periodos_inactivos": periodos_inactivos,
    })


@login_required
@admin_required
def crear_periodo(request):
    """
    Crea un nuevo período en DimPeriodo.
    Validaciones:
    - año y mes obligatorios
    - deben ser numéricos
    - no debe existir ya la combinación (anio, mes)

    Importante:
    - Tu tabla usa id_periodo manual: por eso buscas el último y sumas +1.
    """
    if request.method == "POST":
        anio = request.POST.get("anio")
        mes = request.POST.get("mes")
        descripcion = request.POST.get("descripcion")

        # Validaciones básicas
        if not anio or not mes:
            contexto = {
                "error": "Año y mes son obligatorios.",
                "anio": anio,
                "mes": mes,
                "descripcion": descripcion,
            }
            return render(request, "crear_periodo.html", contexto)

        # Parseo seguro
        try:
            anio_int = int(anio)
            mes_int = int(mes)
        except ValueError:
            contexto = {
                "error": "Año y mes deben ser numéricos.",
                "anio": anio,
                "mes": mes,
                "descripcion": descripcion,
            }
            return render(request, "crear_periodo.html", contexto)

        # Evitar duplicados por clave lógica (anio, mes)
        if DimPeriodo.objects.filter(anio=anio_int, mes=mes_int).exists():
            contexto = {
                "error": f"Ya existe un período con año {anio_int} y mes {mes_int:02d}. "
                         "Edítalo en lugar de crearlo nuevamente.",
                "anio": anio,
                "mes": mes,
                "descripcion": descripcion,
            }
            return render(request, "crear_periodo.html", contexto)

        # Autoincremento manual de id_periodo
        ultimo = DimPeriodo.objects.order_by("-id_periodo").first()
        nuevo_id = (ultimo.id_periodo + 1) if ultimo else 1

        # Crear período
        periodo = DimPeriodo.objects.create(
            id_periodo=nuevo_id,
            anio=anio_int,
            mes=mes_int,
            descripcion=descripcion,
            creado_en=timezone.now(),
        )

        # Auditoría (trazabilidad)
        registrar_auditoria(
            request,
            AuditLog.ACCION_PERIODO,
            f"Creó período {periodo.anio}-{periodo.mes:02d} ({periodo.descripcion}).",
        )

        return redirect("lista_periodos")

    # GET → mostrar formulario
    return render(request, "crear_periodo.html")


@login_required
@admin_required
def editar_periodo(request, id_periodo):
    """
    Edita un período existente.
    Nota: aquí permites cambiar año/mes/descripcion. Mantienes el mismo id_periodo.
    """
    periodo = get_object_or_404(DimPeriodo, pk=id_periodo)

    if request.method == "POST":
        anio = request.POST.get("anio")
        mes = request.POST.get("mes")
        descripcion = request.POST.get("descripcion")

        if not anio or not mes:
            return HttpResponse("Año y mes son obligatorios.")

        try:
            anio = int(anio)
            mes = int(mes)
        except ValueError:
            return HttpResponse("Año y mes deben ser numéricos.")

        # Guardar cambios
        periodo.anio = anio
        periodo.mes = mes
        periodo.descripcion = descripcion
        periodo.save()

        # Auditoría
        registrar_auditoria(
            request,
            AuditLog.ACCION_PERIODO,
            f"Editó período ID {periodo.id_periodo}: {periodo.anio}-{periodo.mes:02d} ({periodo.descripcion}).",
        )

        return redirect("lista_periodos")

    return render(request, "editar_periodo.html", {
        "periodo": periodo
    })


@login_required
@admin_required
def eliminar_periodo(request, id_periodo):
    """
    "Eliminación" lógica:
    - No se borra el registro.
    - Se agrega la marca [INACTIVO] en la descripción.
    Esto evita perder trazabilidad y mantiene consistencia histórica.
    """
    periodo = get_object_or_404(DimPeriodo, pk=id_periodo)

    if request.method == "POST":
        texto = periodo.descripcion or ""
        if "[INACTIVO]" not in texto:
            texto = texto.strip() + " [INACTIVO]"
        periodo.descripcion = texto.strip()
        periodo.save(update_fields=["descripcion"])

        registrar_auditoria(
            request,
            AuditLog.ACCION_PERIODO,
            f"Marcó como INACTIVO el período ID {periodo.id_periodo} ({periodo.anio}-{periodo.mes:02d}).",
        )

        return redirect("lista_periodos")

    return render(request, "confirmar_eliminar_periodo.html", {
        "periodo": periodo,
    })


@login_required
@admin_required
def reactivar_periodo(request, id_periodo):
    """
    Reactiva un período eliminado lógicamente:
    - Quita la marca [INACTIVO] del texto de descripción.
    Se invoca por POST (normalmente con confirmación en UI).
    """
    periodo = get_object_or_404(DimPeriodo, pk=id_periodo)

    if request.method == "POST":
        desc = (periodo.descripcion or "")
        desc = desc.replace("[INACTIVO]", "").strip()
        periodo.descripcion = desc
        periodo.save()

        registrar_auditoria(
            request,
            AuditLog.ACCION_PERIODO,
            f"Reactivó período ID {periodo.id_periodo} ({periodo.anio}-{periodo.mes:02d}).",
        )

        return redirect("lista_periodos")

    return redirect("lista_periodos")


# ========================
# SUBIR Y LISTAR ARCHIVOS
# ========================
@login_required
def subir_excel(request):
    """
    Sube uno o varios archivos REM (Excel) y los asocia a un DimPeriodo.

    Validaciones:
    - extensión .xlsx
    - tamaño máximo 20 MB
    - estructura mínima REM (existencia hoja A01)
    - evita duplicados (mismo nombre + mismo período)

    Resultado:
    - solo archivos válidos se guardan
    - se informa al usuario mediante mensajes claros
    """

    periodos = DimPeriodo.objects.all().order_by("-anio", "-mes")

    if request.method == 'POST':
        archivos = request.FILES.getlist('archivos')
        periodo_id = request.POST.get('periodo_id')

        # ==========================
        # Validaciones iniciales
        # ==========================
        if not archivos:
            messages.error(request, "No seleccionaste ningún archivo.")
            return redirect('subir_excel')

        if not periodo_id:
            messages.error(request, "Debes seleccionar un período.")
            return redirect('subir_excel')

        # ==========================
        # Obtener período
        # ==========================
        try:
            periodo = DimPeriodo.objects.get(pk=periodo_id)
        except DimPeriodo.DoesNotExist:
            messages.error(request, "El período seleccionado no existe.")
            return redirect('subir_excel')

        archivos_guardados = 0

        # ==========================
        # Procesamiento de archivos
        # ==========================
        for archivo in archivos:

            # 0) Evitar duplicados (mismo archivo + mismo período)
            if ArchivoREM.objects.filter(
                periodo=periodo,
                nombre_original=archivo.name
            ).exists():
                messages.warning(
                    request,
                    f"El archivo '{archivo.name}' ya fue cargado previamente en este período."
                )
                continue

            # 1) Validar extensión
            extension = os.path.splitext(archivo.name)[1].lower()
            if extension != '.xlsx':
                messages.warning(
                    request,
                    f"El archivo '{archivo.name}' fue ignorado: formato no válido."
                )
                continue

            # 2) Validar tamaño máximo (20 MB)
            if archivo.size > 20 * 1024 * 1024:
                messages.warning(
                    request,
                    f"El archivo '{archivo.name}' fue ignorado: supera el tamaño máximo permitido."
                )
                continue

            # 3) Validar estructura mínima REM (hoja A01)
            try:
                wb = load_workbook(archivo, read_only=True, data_only=True)
            except Exception:
                messages.warning(
                    request,
                    f"El archivo '{archivo.name}' no pudo ser leído y fue ignorado."
                )
                continue

            if "A01" not in wb.sheetnames:
                messages.warning(
                    request,
                    f"El archivo '{archivo.name}' no corresponde a un REM válido (falta hoja A01)."
                )
                continue

            # IMPORTANTE: resetear puntero del archivo
            archivo.seek(0)

            # 4) Guardar archivo válido
            nuevo = ArchivoREM.objects.create(
                nombre_original=archivo.name,
                archivo=archivo,
                periodo=periodo
            )

            archivos_guardados += 1

            registrar_auditoria(
                request,
                AuditLog.ACCION_UPLOAD,
                f"Subió archivo REM '{nuevo.nombre_original}' "
                f"para el período {periodo.anio}-{periodo.mes:02d}.",
            )

        # ==========================
        # Resultado final
        # ==========================
        if archivos_guardados == 0:
            messages.error(
                request,
                "No se cargó ningún archivo válido. Revisa los mensajes de advertencia."
            )
            return redirect('subir_excel')

        messages.success(
            request,
            f"Se cargaron correctamente {archivos_guardados} archivo(s) REM."
        )
        return redirect('lista_archivos')

    return render(request, 'subir_excel.html', {
        "periodos": periodos,
    })


@login_required
def lista_archivos(request):
    """
    Lista archivos REM activos.
    Nota: excluye "INGRESO MANUAL ..." porque esos se crean internamente
    para registros manuales por período/hoja/sección.
    """
    archivos = (
        ArchivoREM.objects
        .filter(activo=True)
        .exclude(nombre_original__startswith="INGRESO MANUAL ")
        .select_related("periodo")
        .order_by("-fecha_carga")   # campo real de orden
    )

    total_activos = archivos.count()

    return render(request, "lista_archivos.html", {
        "archivos": archivos,
        "total_activos": total_activos,
    })


# ========================
# Procesamiento genérico con mapeo + guardado en BD
# ========================
@login_required
@admin_required
def procesar_archivo_generico(request, archivo_id):
    """
    Procesa un archivo REM completo:
    1) Lee Excel con procesar_archivo_con_mapeo(ruta) -> lista de dicts
    2) Borra registros previos de ese ArchivoREM (evita duplicados)
    3) Inserta masivamente en RegistroREM (bulk_create)
    4) Marca el archivo como procesado
    5) Registra auditoría
    6) Muestra un resumen por hoja/sección

    Nota:
    - Se usa transaction.atomic() para asegurar consistencia al guardar.
    """
    archivo_rem = get_object_or_404(ArchivoREM, pk=archivo_id)
    ruta = archivo_rem.archivo.path

    # 1) Intentar procesar Excel -> lista de diccionarios
    try:
        registros_dict = procesar_archivo_con_mapeo(ruta)
    except Exception as e:
        return HttpResponse(
            f"""
            <h2>Error al procesar el archivo</h2>
            <p><strong>{archivo_rem.nombre_original}</strong></p>
            <p>Detalle técnico del error (para depuración):</p>
            <pre>{str(e)}</pre>
            <a href="/archivos/">⬅ Volver a archivos</a>
            """,
            status=500
        )

    # 2) Borrar registros anteriores de este archivo (evita duplicados)
    RegistroREM.objects.filter(archivo=archivo_rem).delete()

    # 3) Convertir dicts en objetos RegistroREM
    objetos = []
    resumen_hoja_seccion = Counter()

    for reg in registros_dict:
        # Claves de control (no van dentro de "datos")
        hoja = reg.pop("hoja", "")
        seccion = reg.pop("seccion", "")
        fila = reg.pop("fila", 0)

        objetos.append(
            RegistroREM(
                archivo=archivo_rem,
                hoja=hoja,
                seccion=seccion,
                fila=fila,
                datos=reg,
            )
        )

        resumen_hoja_seccion[(hoja, seccion)] += 1

    # 4) Guardar masivamente dentro de una transacción
    try:
        with transaction.atomic():
            if objetos:
                RegistroREM.objects.bulk_create(objetos, batch_size=1000)

            archivo_rem.procesado = True
            archivo_rem.save(update_fields=["procesado"])
    except Exception as e:
        return HttpResponse(
            f"""
            <h2>Error al guardar los registros en la base de datos</h2>
            <p><strong>{archivo_rem.nombre_original}</strong></p>
            <p>Detalle técnico del error (para depuración):</p>
            <pre>{str(e)}</pre>
            <a href="/archivos/">⬅ Volver a archivos</a>
            """,
            status=500
        )

    registrar_auditoria(
        request,
        AuditLog.ACCION_PROCESAR,
        f"Procesó archivo REM '{archivo_rem.nombre_original}' "
        f"({len(objetos)} registros guardados en RegistroREM).",
    )

    # 5) Preparar detalle para vista (resumen por hoja/sección)
    detalle_listado = []
    for (hoja, seccion), cantidad in sorted(resumen_hoja_seccion.items()):
        detalle_listado.append({
            "hoja": hoja,
            "seccion": seccion,
            "cantidad": cantidad,
        })

    # 6) Renderizar template de resultado
    return render(
        request,
        "resultado_procesar_archivo.html",
        {
            "archivo": archivo_rem,
            "total_registros": len(objetos),
            "detalle": detalle_listado,
        }
    )


# ---------------------------------------
# Metadatos de secciones (título por hoja/sección)
# ---------------------------------------
REM_TITULOS_SECCION = {
    ("A02", "A"): "SECCIÓN A: EMP REALIZADO POR PROFESIONAL",
    ("A02", "B"): "SECCIÓN B: EMP SEGÚN RESULTADO DEL ESTADO NUTRICIONAL",
    ("A02", "C"): "SECCIÓN C: RESULTADOS DE EMP SEGÚN ESTADO DE SALUD",
    ("A02", "D"): "SECCIÓN D: RESULTADOS DE EMP SEGÚN ESTADO DE SALUD (EXÁMENES DE LABORATORIO)",
}


def pretty_col_name(col: str) -> str:
    """
    Convierte un nombre técnico (snake_case) a un label legible.

    Caso especial:
    - Si la columna sigue patrón "edad_X_anos_hombres/mujeres", lo transforma a:
      "X años (Hombres/Mujeres)".
    """
    if col.startswith("edad_") and "anos_" in col:
        resto = col[len("edad_"):]
        partes = resto.split("_")
        if "anos" in partes:
            idx_anos = partes.index("anos")
            rango = " ".join(partes[:idx_anos])
            rango = rango.replace(" ", "–") if " " in rango and "y" not in rango else rango.replace(" ", " ")
            sexo = partes[idx_anos + 1] if idx_anos + 1 < len(partes) else ""
            texto = f"{rango} años ({sexo.capitalize()})"
            texto = texto.replace("mas", "más")
            return texto

    texto = col.replace("_", " ")
    texto = texto.replace("anos", "años")
    return texto.capitalize()


@login_required
def ver_registros_archivo(request, archivo_id):
    """
    Muestra registros procesados de un ArchivoREM, con filtros por hoja y sección.
    Además:
    - Si existe estructura en REM_STRUCTURES, se respeta el orden y número de columnas.
    - Si no existe, se "descubre" columnas desde los datos (dim vs num).
    - Se construyen grupos de columnas (Rango etario / Sexo / Identificación de género)
      para emular el header del Excel.
    """
    archivo = get_object_or_404(ArchivoREM, pk=archivo_id)

    # -----------------------
    # 1) Hojas disponibles en este archivo
    # -----------------------
    hojas_disponibles = (
        RegistroREM.objects.filter(archivo=archivo)
        .values_list("hoja", flat=True)
        .distinct()
        .order_by("hoja")
    )

    hoja = request.GET.get("hoja") or ""
    seccion = request.GET.get("seccion") or ""

    # Si no viene hoja seleccionada, tomar la primera disponible
    if not hoja and hojas_disponibles:
        hoja = hojas_disponibles[0]

    # -----------------------
    # 1.1) Secciones disponibles (según hoja actual)
    # -----------------------
    base_secciones_qs = RegistroREM.objects.filter(archivo=archivo)
    if hoja:
        base_secciones_qs = base_secciones_qs.filter(hoja=hoja)

    secciones_disponibles = (
        base_secciones_qs
        .values_list("seccion", flat=True)
        .distinct()
        .order_by("seccion")
    )

    # Si no viene sección seleccionada, tomar la primera disponible
    if not seccion and secciones_disponibles:
        seccion = secciones_disponibles[0]

    # Etiquetas tipo MINSAL para el combo de REM
    opciones_rem = []
    for h in hojas_disponibles:
        etiqueta = REM_TITULOS_HOJA.get(h, f"REM-{h}")
        opciones_rem.append({
            "codigo": h,
            "label": etiqueta,
        })

    # -----------------------
    # 2) Query filtrada por hoja + sección
    # -----------------------
    qs = RegistroREM.objects.filter(archivo=archivo)
    if hoja:
        qs = qs.filter(hoja=hoja)
    if seccion:
        qs = qs.filter(seccion=seccion)

    qs = qs.order_by("id_registro")

    # Paginación (50 filas por página)
    paginator = Paginator(qs, 50)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # -----------------------
    # 3) Estructura fija desde REM_STRUCTURES (si existe)
    # -----------------------
    hoja_key = (hoja or "").strip().upper()
    seccion_key = (seccion or "").strip().upper()

    estructura_hoja = REM_STRUCTURES.get(hoja_key) or {}
    estructura = estructura_hoja.get(seccion_key) or {}

    columnas_config = estructura.get("columnas")
    num_desc_cols = estructura.get("num_desc_cols")
    usa_estructura_fija = bool(columnas_config)

    columnas_dim = []
    columnas_num = []

    if usa_estructura_fija:
        # Si hay estructura fija, se usa tal cual
        columnas = list(columnas_config)
        if num_desc_cols is None:
            num_desc_cols = 0
    else:
        # Descubrir columnas dinámicamente desde los datos
        for reg in qs:
            datos = reg.datos or {}
            for key, value in datos.items():
                if key in columnas_dim or key in columnas_num:
                    continue
                if value is None:
                    continue

                # Heurística: numérico vs descriptivo
                if isinstance(value, (int, float, Decimal)):
                    columnas_num.append(key)
                else:
                    columnas_dim.append(key)

        columnas = columnas_dim + columnas_num
        num_desc_cols = len(columnas_dim)

    # -----------------------
    # 4) Nombres "bonitos" de columnas
    # -----------------------
    columnas_bonitas = [pretty_col_name(c) for c in columnas]

    # -----------------------
    # 5) Grupos para cabecera (simular Excel con encabezados agrupados)
    # -----------------------
    def grupo_columna(nombre_campo: str) -> str:
        n = (nombre_campo or "").lower()

        if n.startswith("rango_etario_"):
            return "Rango etario"
        if n.startswith("sexo_"):
            return "Sexo"
        if n.startswith("identificacion_de_genero_"):
            return "Identificación de género"
        return ""

    grupos_por_col = [grupo_columna(c) for c in columnas]
    tiene_grupos = any(bool(g) for g in grupos_por_col)

    bloques_header = []
    if columnas:
        if tiene_grupos:
            grupo_actual = grupos_por_col[0]
            span = 1
            for g in grupos_por_col[1:]:
                if g == grupo_actual:
                    span += 1
                else:
                    bloques_header.append({"nombre": grupo_actual, "span": span})
                    grupo_actual = g
                    span = 1
            bloques_header.append({"nombre": grupo_actual, "span": span})
        else:
            bloques_header.append({
                "nombre": "Valores registrados",
                "span": len(columnas),
            })

    # -----------------------
    # 6) Construir filas para la tabla (solo página actual)
    # -----------------------
    filas_tabla = []
    for reg in page_obj:
        datos = reg.datos or {}
        fila = {
            "id": reg.id_registro,
            "hoja": reg.hoja,
            "seccion": reg.seccion,
            "fila_excel": reg.fila,
            "celdas": [datos.get(col) for col in columnas],
        }
        filas_tabla.append(fila)

    # -----------------------
    # 7) Títulos bonitos (hoja y sección)
    # -----------------------
    titulo_hoja = REM_TITULOS_HOJA.get(hoja) if hoja else ""
    titulo_seccion = REM_TITULOS_SECCION.get((hoja, seccion)) if hoja and seccion else ""

    # -----------------------
    # 8) Render
    # -----------------------
    return render(request, "ver_registros.html", {
        "archivo": archivo,
        "page_obj": page_obj,
        "hojas_disponibles": hojas_disponibles,
        "secciones_disponibles": secciones_disponibles,
        "hoja_actual": hoja,
        "seccion_actual": seccion,
        "columnas": columnas,
        "columnas_bonitas": columnas_bonitas,
        "filas_tabla": filas_tabla,
        "titulo_hoja": titulo_hoja,
        "titulo_seccion": titulo_seccion,
        "num_desc_cols": num_desc_cols,
        "usa_estructura_fija": usa_estructura_fija,
        "bloques_header": bloques_header,
        "opciones_rem": opciones_rem,
    })


@login_required
@admin_required
def desactivar_archivo(request, archivo_id):
    """
    "Eliminación" lógica para archivos:
    - Se marca activo=False
    - Se agrega [ANULADO] al nombre_original si no existe
    """
    archivo = get_object_or_404(ArchivoREM, pk=archivo_id)

    if request.method == "POST":
        archivo.activo = False
        if "[ANULADO]" not in archivo.nombre_original:
            archivo.nombre_original = archivo.nombre_original + " [ANULADO]"
        archivo.save(update_fields=["activo", "nombre_original"])

        registrar_auditoria(
            request,
            AuditLog.ACCION_UPLOAD,
            f"Marcó como ANULADO el archivo REM '{archivo.nombre_original}'.",
        )

        return redirect("lista_archivos")

    return render(request, "confirmar_desactivar_archivo.html", {
        "archivo": archivo,
    })


# ========================
# BITÁCORA DE AUDITORÍA
# ========================
@login_required
@admin_required
def lista_auditoria(request):
    """
    Bitácora de auditoría del sistema.
    Permite filtros por:
    - usuario (contiene texto)
    - acción (exacta)
    """
    logs = AuditLog.objects.select_related("usuario").all()

    usuario = request.GET.get("usuario") or ""
    accion = request.GET.get("accion") or ""

    if usuario:
        logs = logs.filter(usuario__username__icontains=usuario)
    if accion:
        logs = logs.filter(accion=accion)

    logs = logs.order_by("-fecha_hora")

    paginator = Paginator(logs, 50)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    acciones = AuditLog.ACCIONES_CHOICES

    return render(request, "lista_auditoria.html", {
        "page_obj": page_obj,
        "acciones": acciones,
        "filtro_usuario": usuario,
        "filtro_accion": accion,
    })


@login_required
@admin_required
def ingresar_registro_manual_periodo(request, periodo_id, hoja, seccion):
    """
    Permite ingreso manual (sin Excel real) por período + REM + sección.

    - Si ya existen registros manuales, el formulario NO precarga datos
      (se visualizan en reportes), pero se informa al usuario.
    """

    periodo = get_object_or_404(DimPeriodo, id_periodo=periodo_id)

    hoja_key_up = (hoja or "").strip().upper()
    seccion_key_up = (seccion or "").strip().upper()

    # -----------------------------
    # Estructura desde REM_STRUCTURES
    # -----------------------------
    estructura_hoja = (
        REM_STRUCTURES.get(hoja_key_up)
        or REM_STRUCTURES.get(hoja_key_up.lower())
        or {}
    )
    estructura = (
        estructura_hoja.get(seccion_key_up)
        or estructura_hoja.get(seccion_key_up.lower())
        or {}
    )

    columnas = estructura.get("columnas")
    if not columnas:
        return HttpResponse(
            f"No hay estructura configurada para {hoja_key_up} / sección {seccion_key_up}.",
            status=400,
        )

    # -----------------------------
    # Filas fijas SOLO para REM A01 / Sección A
    # -----------------------------
    filas_fijas = []
    if hoja_key_up == "A01" and seccion_key_up == "A":
        filas_fijas = [
            ("Preconcepcional", "Médico/a"),
            ("Preconcepcional", "Matrona/ón"),
            ("Prenatal", "Médico/a"),
            ("Prenatal", "Matrona/ón"),
            ("Post Parto", "Médico/a"),
            ("Post Parto", "Matrona/ón"),
            ("Post Aborto", "Médico/a"),
            ("Post Aborto", "Matrona/ón"),
            ("Púerpera con RN hasta 10 días", "Médico/a"),
            ("Púerpera con RN hasta 10 días", "Matrona/ón"),
            ("Púerpera con RN entre 11 y 28 días", "Médico/a"),
            ("Púerpera con RN entre 11 y 28 días", "Matrona/ón"),
            ("RN hasta 10 días de vida", "Médico/a"),
            ("RN hasta 10 días de vida", "Matrona/ón"),
            ("RN entre 11 y 28 días de vida", "Médico/a"),
            ("RN entre 11 y 28 días de vida", "Matrona/ón"),
            ("Ginecológico", "Médico/a"),
            ("Ginecológico", "Matrona/ón"),
            ("Climaterio", "Médico/a"),
            ("Climaterio", "Matrona/ón"),
            ("Regulación de Fecundidad", "Médico/a"),
            ("Regulación de Fecundidad", "Matrona/ón"),
        ]

    # -----------------------------
    # ArchivoREM MANUAL asociado
    # -----------------------------
    nombre_manual = (
        f"INGRESO MANUAL {hoja_key_up}-{seccion_key_up} "
        f"{periodo.anio}-{periodo.mes:02d}"
    )

    archivo_manual, creado = ArchivoREM.objects.get_or_create(
        periodo=periodo,
        nombre_original=nombre_manual,
        defaults={
            "archivo": (
                f"rem_uploads/manual_{hoja_key_up}_{seccion_key_up}_"
                f"{periodo.anio}_{periodo.mes:02d}.xlsx"
            ),
            "procesado": True,
            "activo": True,
        },
    )

    # ======================================================
    # DETECTAR SI YA EXISTEN REGISTROS MANUALES (CLAVE)
    # ======================================================
    registros_existentes = RegistroREM.objects.filter(
        archivo=archivo_manual,
        hoja=hoja_key_up,
        seccion=seccion_key_up,
    )
    hay_datos = registros_existentes.exists()

    # -----------------------------
    # POST: guardar registros
    # -----------------------------
    if request.method == "POST":
        registros_a_crear = []

        ultimo = (
            RegistroREM.objects
            .filter(archivo=archivo_manual, hoja=hoja_key_up, seccion=seccion_key_up)
            .order_by("-fila")
            .first()
        )
        siguiente_fila = (ultimo.fila + 1) if ultimo else 1

        if filas_fijas:
            for idx, (tipo_control, profesional) in enumerate(filas_fijas):
                datos = {}

                for col in columnas:
                    if col == "tipo_de_control":
                        datos[col] = tipo_control
                        continue
                    if col == "profesional":
                        datos[col] = profesional
                        continue

                    key = f"{col}__{idx}"
                    raw = (request.POST.get(key, "") or "").strip()

                    if raw == "":
                        datos[col] = None
                    else:
                        try:
                            datos[col] = int(raw)
                        except ValueError:
                            try:
                                datos[col] = float(raw.replace(",", "."))
                            except ValueError:
                                datos[col] = raw

                hay_valores = any(
                    (v not in (None, "", 0))
                    for k, v in datos.items()
                    if k not in ("tipo_de_control", "profesional")
                )
                if not hay_valores:
                    continue

                registros_a_crear.append(
                    RegistroREM(
                        archivo=archivo_manual,
                        hoja=hoja_key_up,
                        seccion=seccion_key_up,
                        fila=siguiente_fila,
                        datos=datos,
                    )
                )
                siguiente_fila += 1
        else:
            datos = {}
            for col in columnas:
                raw = (request.POST.get(col, "") or "").strip()
                if raw == "":
                    datos[col] = None
                else:
                    try:
                        datos[col] = int(raw)
                    except ValueError:
                        try:
                            datos[col] = float(raw.replace(",", "."))
                        except ValueError:
                            datos[col] = raw

            registros_a_crear.append(
                RegistroREM(
                    archivo=archivo_manual,
                    hoja=hoja_key_up,
                    seccion=seccion_key_up,
                    fila=siguiente_fila,
                    datos=datos,
                )
            )

        if registros_a_crear:
            RegistroREM.objects.bulk_create(registros_a_crear)

            registrar_auditoria(
                request,
                AuditLog.ACCION_OTRA,
                (
                    f"Ingreso manual de {len(registros_a_crear)} registro(s) en "
                    f"{hoja_key_up}-{seccion_key_up} para el período "
                    f"{periodo.anio}-{periodo.mes:02d} "
                    f"(ArchivoREM ID {archivo_manual.id_archivo})."
                ),
            )

        return redirect("ver_registros_archivo", archivo_id=archivo_manual.id_archivo)

    # -----------------------------
    # GET: mostrar formulario
    # -----------------------------
    contexto = {
        "periodo": periodo,
        "hoja": hoja_key_up,
        "seccion": seccion_key_up,
        "columnas": columnas,
        "filas_fijas": filas_fijas,
        "hay_datos": hay_datos,  # 👈 CLAVE
    }
    return render(request, "ingresar_registro_manual.html", contexto)


@login_required
def seleccionar_seccion_periodo(request, periodo_id, hoja):
    """
    Pantalla 2 (flujo por período):
    - Recibe periodo_id + hoja (REM)
    - Lista secciones disponibles según REM_STRUCTURES
    - Muestra cuántas columnas tiene cada sección (indicativo de "tamaño")
    """
    periodo = get_object_or_404(DimPeriodo, id_periodo=periodo_id)

    hoja_key_up = (hoja or "").strip().upper()
    rem_data = (
        REM_STRUCTURES.get(hoja_key_up)
        or REM_STRUCTURES.get(hoja_key_up.lower())
    )
    if not rem_data or not isinstance(rem_data, dict):
        return HttpResponse(f"REM {hoja_key_up} no encontrado.", status=404)

    # REM_STRUCTURES puede venir con "secciones" o directo
    if "secciones" in rem_data and isinstance(rem_data["secciones"], dict):
        secciones_dict = rem_data["secciones"]
    else:
        secciones_dict = {
            k: v for k, v in rem_data.items()
            if isinstance(v, dict)
        }

    secciones = []
    for sec_key, sec_data in secciones_dict.items():
        seccion = str(sec_key).upper()
        columnas = sec_data.get("columnas") or []
        titulo_seccion = sec_data.get("titulo", "")
        if not columnas:
            continue

        secciones.append({
            "seccion": seccion,
            "titulo": titulo_seccion,
            "num_columnas": len(columnas),
        })

    secciones = sorted(secciones, key=lambda s: s["seccion"])

    contexto = {
        "periodo": periodo,
        "hoja": hoja_key_up,
        "secciones": secciones,
    }
    return render(request, "seleccionar_seccion_periodo.html", contexto)


@login_required
def seleccionar_rem_periodo(request, periodo_id):
    """
    Pantalla 1 (flujo por período):
    - Lista REM disponibles (A01, A02, A03...) según REM_STRUCTURES.
    """
    periodo = get_object_or_404(DimPeriodo, id_periodo=periodo_id)

    rems = []
    for rem_key, rem_data in REM_STRUCTURES.items():
        if not isinstance(rem_data, dict):
            continue

        hoja = str(rem_key).upper()
        nombre_rem = rem_data.get("nombre") or REM_TITULOS_HOJA.get(hoja, hoja)

        rems.append({
            "hoja": hoja,
            "nombre": nombre_rem,
        })

    rems = sorted(rems, key=lambda r: r["hoja"])

    return render(
        request,
        "seleccionar_rem_periodo.html",
        {"periodo": periodo, "rems": rems},
    )


@login_required
def ver_datos_rem_periodo(request, periodo_id):
    """
    Resumen por período:
    - Agrupa por hoja y sección
    - Muestra total de registros por sección

    Fuente:
    - RegistroREM filtrando por ArchivoREM.periodo
    """
    periodo = get_object_or_404(DimPeriodo, id_periodo=periodo_id)

    resumen_qs = (
        RegistroREM.objects
        .filter(archivo__periodo=periodo)
        .values('hoja', 'seccion')
        .annotate(total_registros=Count('id_registro'))
        .order_by('hoja', 'seccion')
    )

    grupos_dict = defaultdict(list)
    for r in resumen_qs:
        grupos_dict[r['hoja']].append(r)

    grupos = []
    for hoja, secciones in sorted(grupos_dict.items()):
        grupos.append({
            "hoja": hoja,
            "nombre": REM_TITULOS_HOJA.get(hoja, f"REM-{hoja}"),
            "secciones": secciones,
        })

    return render(
        request,
        "ver_datos_rem_periodo.html",
        {
            "periodo": periodo,
            "grupos": grupos,
        }
    )


@login_required
def ver_detalle_rem(request, periodo_id, hoja, seccion):
    """
    Detalle por período + hoja + sección:
    - Muestra tabla completa (sin paginar aquí)
    - Usa la misma lógica de columnas (estructura fija o dinámica)

    Nota:
    - La PK del período es id_periodo (no "id").
    """
    periodo = get_object_or_404(DimPeriodo, id_periodo=periodo_id)

    hoja = (hoja or "").strip().upper()
    seccion = (seccion or "").strip().upper()

    # -----------------------
    # 1) Query filtrada por período + hoja + sección
    # -----------------------
    qs = (
        RegistroREM.objects
        .filter(archivo__periodo=periodo, hoja=hoja, seccion=seccion)
        .order_by("id_registro")
    )

    # -----------------------
    # 2) Columnas por REM_STRUCTURES (si existe)
    # -----------------------
    hoja_key = hoja
    seccion_key = seccion

    estructura_hoja = REM_STRUCTURES.get(hoja_key) or {}
    estructura = estructura_hoja.get(seccion_key) or {}

    columnas_config = estructura.get("columnas")
    num_desc_cols = estructura.get("num_desc_cols")
    usa_estructura_fija = bool(columnas_config)

    columnas_dim = []
    columnas_num = []

    if usa_estructura_fija:
        columnas = list(columnas_config)
        if num_desc_cols is None:
            num_desc_cols = 0
    else:
        for reg in qs:
            datos = reg.datos or {}
            for key, value in datos.items():
                if key in columnas_dim or key in columnas_num:
                    continue
                if value is None:
                    continue

                if isinstance(value, (int, float, Decimal)):
                    columnas_num.append(key)
                else:
                    columnas_dim.append(key)

        columnas = columnas_dim + columnas_num
        num_desc_cols = len(columnas_dim)

    # -----------------------
    # 3) Encabezados legibles
    # -----------------------
    columnas_bonitas = [pretty_col_name(c) for c in columnas]

    # -----------------------
    # 4) Filas tabla
    # -----------------------
    filas_tabla = []
    for reg in qs:
        datos = reg.datos or {}
        fila = {
            "id": reg.id_registro,
            "hoja": reg.hoja,
            "seccion": reg.seccion,
            "fila_excel": reg.fila,
            "celdas": [datos.get(col) for col in columnas],
            "es_total": False,
        }
        filas_tabla.append(fila)

    # -----------------------
    # 5) Títulos bonitos
    # -----------------------
    titulo_hoja = REM_TITULOS_HOJA.get(hoja) if hoja else ""
    titulo_seccion = REM_TITULOS_SECCION.get((hoja, seccion)) if hoja and seccion else ""

    # -----------------------
    # 6) Render
    # -----------------------
    context = {
        "periodo": periodo,
        "hoja": hoja,
        "seccion": seccion,
        "columnas_bonitas": columnas_bonitas,
        "filas_tabla": filas_tabla,
        "num_desc_cols": num_desc_cols,
        "titulo_hoja": titulo_hoja,
        "titulo_seccion": titulo_seccion,
    }
    return render(request, "ver_detalle_rem.html", context)


@login_required
def reportes_home(request):
    """
    Home de reportes:

    - Lista SOLO períodos activos (excluye [INACTIVO])
    - Muestra estado de consolidación por período
    - NO oculta períodos sin archivos
    - Explica por qué un período no tiene reportes disponibles

    Estados posibles:
    - sin_archivos
    - cargado_no_procesado
    - parcial
    - completo
    """

    # ==============================
    # 1) Obtener solo períodos activos
    # ==============================
    periodos = (
        DimPeriodo.objects
        .exclude(descripcion__icontains='[INACTIVO]')
        .order_by("-anio", "-mes")
    )

    estado_periodos = []

    # ==============================
    # 2) Calcular estado por período
    # ==============================
    for p in periodos:
        archivos_qs = ArchivoREM.objects.filter(periodo=p, activo=True)

        total_archivos = archivos_qs.count()
        procesados = archivos_qs.filter(procesado=True).count()

        # -----------------------------
        # Determinar estado del período
        # -----------------------------
        if total_archivos == 0:
            estado = "sin_archivos"
            porcentaje = 0

        elif procesados == 0:
            estado = "cargado_no_procesado"
            porcentaje = 0

        elif procesados < total_archivos:
            estado = "parcial"
            porcentaje = int((procesados / total_archivos) * 100)

        else:
            estado = "completo"
            porcentaje = 100

        estado_periodos.append({
            "periodo": p,
            "total_archivos": total_archivos,
            "procesados": procesados,
            "porcentaje": porcentaje,
            "estado": estado,
        })

    # ==============================
    # 3) Renderizar vista
    # ==============================
    return render(request, "reportes_home.html", {
        "estado_periodos": estado_periodos,
    })


@login_required
def reporte_a01_seccion_a(request, periodo_id):
    """
    Reporte específico: REM A01 - Sección A (por período)
    - Consulta registros procesados (archivos activos)
    - Calcula resumen y datos para gráficos (tipos / profesionales / rangos)
    """
    periodo = get_object_or_404(DimPeriodo, pk=periodo_id)

    registros = RegistroREM.objects.filter(
        archivo__periodo=periodo,
        archivo__activo=True,
        hoja="A01",
        seccion="A",
    )

    # Resumen (KPIs + alertas)
    resumen = calcular_resumen_a01_seccion_a(registros, periodo)

    # Acumuladores para gráficos
    tipos_counter = Counter()
    prof_counter = Counter()
    rangos_counter = Counter()

    for r in registros:
        d = r.datos or {}

        # TOTAL REAL = suma de rangos etarios (no confía ciegamente en "total")
        total_fila = 0
        for k, v in d.items():
            if k.startswith("rango_etario_"):
                total_fila += _to_int(v)

        if total_fila <= 0:
            continue

        tipo = (d.get("tipo_de_control") or "").strip() or "Otros controles"
        profesional = (d.get("profesional") or "").strip() or "Sin profesional"

        tipos_counter[tipo] += total_fila
        prof_counter[profesional] += total_fila

        for k, v in d.items():
            if k.startswith("rango_etario_"):
                rangos_counter[k] += _to_int(v)

    # Helper para armar "top N" con porcentajes
    def build_chart(counter, labels_map=None, max_items=6):
        items = [(k, v) for k, v in counter.items() if v > 0]
        if not items:
            return []

        items.sort(key=lambda kv: kv[1], reverse=True)
        items = items[:max_items]

        total = sum(v for _, v in items) or 1
        data = []
        for key, valor in items:
            nombre = labels_map.get(key, key) if labels_map else key
            pct = round(valor * 100 / total)
            data.append({
                "nombre": nombre,
                "valor": valor,
                "porcentaje": pct,
            })
        return data

    # Labels legibles para rangos
    RANGO_LABELS = {
        "rango_etario_menos_de_4_anos": "Menos de 4 años",
        "rango_etario_5_9_anos": "5 - 9 años",
        "rango_etario_10_14_anos": "10 - 14 años",
        "rango_etario_15_19_anos": "15 - 19 años",
        "rango_etario_20_24_anos": "20 - 24 años",
        "rango_etario_25_29_anos": "25 - 29 años",
        "rango_etario_30_34_anos": "30 - 34 años",
        "rango_etario_35_39_anos": "35 - 39 años",
        "rango_etario_40_44_anos": "40 - 44 años",
        "rango_etario_45_49_anos": "45 - 49 años",
        "rango_etario_50_54_anos": "50 - 54 años",
        "rango_etario_55_59_anos": "55 - 59 años",
        "rango_etario_60_64_anos": "60 - 64 años",
        "rango_etario_65_69_anos": "65 - 69 años",
        "rango_etario_70_74_anos": "70 - 74 años",
        "rango_etario_75_79_anos": "75 - 79 años",
        "rango_etario_80_y_mas_anos": "80 y más años",
    }

    tipos_chart = build_chart(tipos_counter)
    profesionales_chart = build_chart(prof_counter)
    rangos_chart = build_chart(rangos_counter, RANGO_LABELS)

    return render(request, "reporte_a01_seccion_a.html", {
        "periodo": periodo,
        "resumen": resumen,
        "tipos_chart": tipos_chart,
        "profesionales_chart": profesionales_chart,
        "rangos_chart": rangos_chart,
    })


# ========================
# EXPORTAR (REM A01 - SECCIÓN A)
# ========================
@login_required
@admin_required
def exportar_a01_seccion_a_excel(request, periodo_id):
    """
    Exporta a Excel la tabla REM A01 - Sección A para un período.

    Implementación:
    - Construye una cabecera de 2 filas con merges
    - Luego escribe datos desde fila 3, respetando el orden A..AF
    """
    periodo = get_object_or_404(DimPeriodo, pk=periodo_id)

    registros = RegistroREM.objects.filter(
        archivo__periodo=periodo,
        archivo__activo=True,
        hoja="A01",
        seccion="A",
    ).order_by("id_registro")

    wb = Workbook()
    ws = wb.active
    ws.title = "REM A01 - Sección A"

    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # --------------------------
    # 1) CABECERA DE DOS FILAS
    # --------------------------
    ws["A1"] = "TIPO DE CONTROL"
    ws["B1"] = "PROFESIONAL"
    ws["C1"] = "TOTAL"
    ws["D1"] = "RANGO ETARIO"
    ws["U1"] = "SEXO"
    ws["V1"] = "Control con pareja, familiar u otro"
    ws["W1"] = "Control de diada con presencia del padre"
    ws["X1"] = "Espacios Amigables/ Adolescentes"
    ws["Y1"] = "Niños, Niñas, Adolescentes y Jóvenes SENAME"
    ws["Z1"] = "Niños, Niñas, Adolescentes y Jóvenes Mejor Niñez"
    ws["AA1"] = "Pueblos Originarios"
    ws["AB1"] = "Migrantes"
    ws["AC1"] = "Personas con discapacidad"
    ws["AD1"] = "Identificación de género"
    ws["AE1"] = "Adolescente acude a control MAC con pareja"

    # Fila 2: rangos etarios + sexo + identidad de género
    rangos = [
        "Menos de 4 años",
        "5 - 9 años",
        "10 - 14 años",
        "15 - 19 años",
        "20 - 24 años",
        "25 - 29 años",
        "30 - 34 años",
        "35 - 39 años",
        "40 - 44 años",
        "45 - 49 años",
        "50 - 54 años",
        "55 - 59 años",
        "60 - 64 años",
        "65 - 69 años",
        "70 - 74 años",
        "75 - 79 años",
        "80 y más años",
    ]
    col = 4  # D
    for etiqueta in rangos:
        ws.cell(row=2, column=col, value=etiqueta)
        col += 1

    ws["U2"] = "Hombres"
    ws["V2"] = "Mujeres"
    ws["AD2"] = "Trans masculino"
    ws["AE2"] = "Trans femenina"
    ws["AF2"] = "No binarie"

    # --------------------------
    # 2) MERGE CELLS (rowspan / colspan)
    # --------------------------
    ws.merge_cells("A1:A2")
    ws.merge_cells("B1:B2")
    ws.merge_cells("C1:C2")

    ws.merge_cells("D1:T1")   # RANGO ETARIO
    ws.merge_cells("U1:V1")   # SEXO

    # Columnas con rowspan=2
    for col_letter in ["V", "W", "X", "Y", "Z", "AA", "AB", "AC", "AE"]:
        ws.merge_cells(f"{col_letter}1:{col_letter}2")

    ws.merge_cells("AD1:AF1")  # Identificación de género

    # Estilo cabecera
    for row in (1, 2):
        for col_idx in range(1, 33):  # hasta AF aprox
            cell = ws.cell(row=row, column=col_idx)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )
            cell.font = Font(bold=True)
            cell.border = border

    # Ancho columnas
    for col_idx in range(1, 33):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12

    # --------------------------
    # 3) DATOS (DESDE FILA 3)
    # --------------------------
    fila_excel = 3
    for reg in registros:
        d = reg.datos or {}

        # Orden A..AF (debe coincidir con la cabecera)
        valores = [
            d.get("tipo_de_control"),
            d.get("profesional"),
            _to_int(d.get("total")),
        ]

        for key, _etq in RANGOS_A01_A:
            valores.append(_to_int(d.get(key)))

        valores.append(_to_int(d.get("sexo_hombres")))
        valores.append(_to_int(d.get("sexo_mujeres")))

        valores.append(_to_int(d.get("control_con_pareja_familiar_u_otro")))
        valores.append(_to_int(d.get("control_de_diada_con_presencia_del_padre")))
        valores.append(_to_int(d.get("espacios_amigables_adolescentes")))
        valores.append(_to_int(d.get("nna_sename")))
        valores.append(_to_int(d.get("nna_mejor_ninez")))
        valores.append(_to_int(d.get("pueblos_originarios")))
        valores.append(_to_int(d.get("migrantes")))
        valores.append(_to_int(d.get("personas_con_discapacidad")))

        valores.append(_to_int(d.get("identificacion_de_genero_trans_masculino")))
        valores.append(_to_int(d.get("identificacion_de_genero_trans_femenina")))
        valores.append(_to_int(d.get("identificacion_de_genero_no_binarie")))

        valores.append(_to_int(d.get("adolescente_acude_a_control_mac_con_pareja")))

        for col_idx, valor in enumerate(valores, start=1):
            cell = ws.cell(row=fila_excel, column=col_idx, value=valor)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        fila_excel += 1

    filename = f"rem_a01_seccion_a_{periodo.anio}_{periodo.mes:02d}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename=\"{filename}\"'
    wb.save(response)
    return response


@login_required
@admin_required
def exportar_a01_seccion_a_pdf(request, periodo_id):
    """
    Exporta a PDF la tabla REM A01 - Sección A para un período.

    - Usa ReportLab
    - Columnas se toman desde REM_STRUCTURES , para mantener consistencia
    - Repite la fila de headers en cada página (repeatRows=1)
    """
    periodo = get_object_or_404(DimPeriodo, pk=periodo_id)

    registros = RegistroREM.objects.filter(
        archivo__periodo=periodo,
        archivo__activo=True,
        hoja="A01",
        seccion="A",
    ).order_by("id_registro")

    # -----------------------
    # 1) Columnas desde REM_STRUCTURES
    # -----------------------
    rem_a01 = (
        REM_STRUCTURES.get("A01")
        or REM_STRUCTURES.get("a01")
        or {}
    )

    if "secciones" in rem_a01 and isinstance(rem_a01["secciones"], dict):
        secciones_dict = rem_a01["secciones"]
        seccion_data = secciones_dict.get("A") or secciones_dict.get("a") or {}
    else:
        seccion_data = rem_a01.get("A") or rem_a01.get("a") or {}

    columnas = seccion_data.get("columnas") or []
    num_desc_cols = seccion_data.get("num_desc_cols", 2)

    # Fallback si no hay estructura
    if not columnas:
        columnas = ["tipo_de_control", "profesional", "total"] + [k for k, _ in RANGOS_A01_A]
        num_desc_cols = 2

    # -----------------------
    # 2) Encabezados legibles
    # -----------------------
    headers = []
    for col in columnas:
        if col in HEADERS_A01_A:
            headers.append(HEADERS_A01_A[col])
        elif col in RANGOS_A01_A_DICT:
            headers.append(RANGOS_A01_A_DICT[col])
        else:
            nombre = col.replace("_", " ").replace("anos", "años")
            headers.append(nombre.capitalize())

    data = [headers]

    # -----------------------
    # 3) Filas
    # -----------------------
    for r in registros:
        d = r.datos or {}
        fila = []
        for idx, col in enumerate(columnas):
            valor = d.get(col, "")

            # Primeras columnas descriptivas como texto; el resto a int
            if idx < num_desc_cols:
                fila.append(valor or "")
            else:
                fila.append(_to_int(valor))
        data.append(fila)

    if len(data) == 1:
        data.append(["Sin datos para este período"] + [""] * (len(headers) - 1))

    # -----------------------
    # 4) Generar el PDF
    # -----------------------
    buffer = BytesIO()
    page_size = landscape(A4)
    page_width, page_height = page_size
    left_margin = 15
    right_margin = 15
    available_width = page_width - left_margin - right_margin

    doc = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        leftMargin=left_margin,
        rightMargin=right_margin,
        topMargin=20,
        bottomMargin=20,
    )

    num_cols = len(headers)

    # Anchos: las 3 primeras columnas más anchas, resto se reparte
    ancho_tipo = 110
    ancho_prof = 70
    ancho_total = 45

    ancho_restante = available_width - (ancho_tipo + ancho_prof + ancho_total)
    num_restantes = max(num_cols - 3, 1)
    ancho_otros = max(ancho_restante / num_restantes, 20)

    col_widths = []
    for idx in range(num_cols):
        if idx == 0:
            col_widths.append(ancho_tipo)
        elif idx == 1:
            col_widths.append(ancho_prof)
        elif idx == 2:
            col_widths.append(ancho_total)
        else:
            col_widths.append(ancho_otros)

    table = Table(data, colWidths=col_widths, repeatRows=1)

    style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e5e7eb")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),

        ("FONTSIZE", (0, 0), (-1, -1), 6),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (num_desc_cols, 1), (-1, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#fffce8")]),
    ])

    table.setStyle(style)

    elements = [table]
    doc.build(elements)

    pdf = buffer.getvalue()
    buffer.close()

    filename = f"rem_a01_seccion_a_{periodo.anio}_{periodo.mes:02d}.pdf"
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    response.write(pdf)
    return response


# ============================================================
# HELPERS
# ============================================================
def _to_int(valor):
    """
    Normaliza valores numéricos que pueden venir como:
    - None
    - int/float
    - string ("", "-", "12", "12,0", "12.0")

    Retorna siempre int (o 0 si no es convertible).
    """
    if valor is None:
        return 0
    if isinstance(valor, (int, float)):
        return int(valor)
    if isinstance(valor, str):
        valor = valor.strip()
        if not valor or valor == "-":
            return 0
        try:
            return int(valor)
        except ValueError:
            try:
                return int(float(valor.replace(",", ".")))
            except ValueError:
                return 0
    return 0


def calcular_resumen_a01_seccion_a(registros_qs, periodo=None):
    """
    Calcula indicadores y alertas para el reporte REM A01 - Sección A.

    Salidas:
    - total_controles
    - num_tipos_control
    - num_profesionales
    - rango_top
    - rangos_chart (lista label/value)
    - solo_estructura (si todo viene en 0)
    - alertas (mensajes para UI)
    """
    total_controles = 0
    tipos_control = set()
    profesionales = set()
    totales_rangos = {key: 0 for key, _ in RANGOS_A01_A}
    alertas = []
    filas_con_inconsistencia = 0

    for reg in registros_qs:
        datos = reg.datos or {}

        # Total declarado (puede no calzar con rangos)
        total_fila = _to_int(datos.get("total", 0))
        total_controles += total_fila

        tc = (datos.get("tipo_de_control") or "").strip()
        if tc:
            tipos_control.add(tc)

        prof = (datos.get("profesional") or "").strip()
        if prof:
            profesionales.add(prof)

        # Suma rangos para detectar inconsistencias
        suma_rangos_fila = 0
        for key, _ in RANGOS_A01_A:
            v = _to_int(datos.get(key, 0))
            totales_rangos[key] += v
            suma_rangos_fila += v

        # Si hay actividad, validar consistencia total vs suma rangos
        if total_fila > 0 or suma_rangos_fila > 0:
            if total_fila != suma_rangos_fila:
                filas_con_inconsistencia += 1

    # Rango con mayor valor
    rango_top = None
    max_val = 0
    for key, label in RANGOS_A01_A:
        val = totales_rangos[key]
        if val > max_val:
            max_val = val
            rango_top = label
    if max_val == 0:
        rango_top = "Sin datos registrados"

    # Alertas por estructura vacía o rangos sin registros
    suma_todos_rangos = sum(totales_rangos.values())
    solo_estructura = (total_controles == 0 and suma_todos_rangos == 0)
    if solo_estructura:
        alertas.append("REM A01 Sección A tiene solo estructura sin datos (todos los valores son 0).")
    else:
        for key, label in RANGOS_A01_A:
            if totales_rangos[key] == 0:
                alertas.append(f"No hay registros en el rango etario «{label}» en este período.")

    # Inconsistencias detectadas
    if filas_con_inconsistencia > 0:
        alertas.append(
            f"Se detectaron {filas_con_inconsistencia} fila(s) donde el TOTAL no coincide con la suma de rangos etarios."
        )

    # Alerta de plazo (cierre estimado: día 10 del mes siguiente)
    alerta_plazo = None
    if periodo is not None:
        hoy = timezone.now().date()
        anio = periodo.anio
        mes = periodo.mes
        if mes == 12:
            cierre = date(anio + 1, 1, 10)
        else:
            cierre = date(anio, mes + 1, 10)
        dias_restantes = (cierre - hoy).days
        if dias_restantes < 0:
            alerta_plazo = f"El plazo estimado de consolidación para este período venció hace {abs(dias_restantes)} día(s)."
        elif dias_restantes <= 5:
            alerta_plazo = f"Quedan {dias_restantes} día(s) para el cierre estimado de consolidación del período."

    if alerta_plazo:
        alertas.insert(0, alerta_plazo)

    # Datos para gráfico de rangos
    rangos_chart = []
    for key, label in RANGOS_A01_A:
        rangos_chart.append({"label": label, "value": totales_rangos[key]})

    return {
        "total_controles": total_controles,
        "num_tipos_control": len(tipos_control),
        "num_profesionales": len(profesionales),
        "rango_top": rango_top,
        "rangos_chart": rangos_chart,
        "solo_estructura": solo_estructura,
        "alertas": alertas,
    }
