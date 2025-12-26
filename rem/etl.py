import os
import csv
import re

from django.conf import settings
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# ==========================
# Cargar mapeo desde CSV
# ==========================

def cargar_mapeo():
    """
    Lee mapeo_rem.csv y deja todo en un diccionario:
    clave = (hoja, seccion, columna_excel)
    valor = dict con encabezados, campo_destino, tipo_dato, etc.
    """
    ruta_csv = os.path.join(
        settings.BASE_DIR, "rem", "mapeo_rem.csv"
    )
    mapeo = {}

    with open(ruta_csv, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            hoja = row["hoja"].strip().upper()
            seccion = row["seccion"].strip().upper()
            col = row["columna_excel"].strip().upper()

            clave = (hoja, seccion, col)
            mapeo[clave] = row

    return mapeo


MAPEO = None  # se inicializa lazy


def get_mapeo():
    global MAPEO
    if MAPEO is None:
        MAPEO = cargar_mapeo()
    return MAPEO


# ==========================
# Helpers para leer secciones del Excel
# ==========================

def limpiar_texto(texto):
    if texto is None:
        return ""
    return str(texto).strip()


def es_fila_vacia(row):
    return all(c is None or str(c).strip() == "" for c in row)


def es_fila_header(row):
    textos = 0
    total = 0
    for c in row:
        if c is None:
            continue
        valor = str(c).strip()
        if not valor:
            continue
        total += 1
        if re.fullmatch(r"[0-9]+", valor):
            continue
        textos += 1
    return textos > 0 and total > 0


# ==========================================
# NUEVO → Normalización de SECCIONES
# ==========================================

def normalizar_seccion(id_sec):
    """
    Normaliza casos como:
    - "G." → "G"
    - "H." → "H"
    - "D1" → "D.1"
    - "A2" → "A.2"
    - "B10" → "B.10"
    """

    if not id_sec:
        return id_sec

    id_sec = id_sec.strip().upper()

    # Caso G. → G
    if re.fullmatch(r"[A-Z]\.", id_sec):
        return id_sec[0]

    # Caso D1, A2, B3 → D.1, A.2, B.3
    if re.fullmatch(r"[A-Z][0-9]", id_sec):
        return f"{id_sec[0]}.{id_sec[1]}"

    # Caso B12 → B.12
    if re.fullmatch(r"[A-Z][0-9]{2,}", id_sec):
        letra = id_sec[0]
        nums = id_sec[1:]
        return f"{letra}.{nums}"

    return id_sec


def extraer_secciones_de_hoja(ws):
    """
    Encuentra las filas donde dice 'SECCIÓN X' y devuelve [{"fila_titulo", "id_seccion"}]
    """
    secciones = []
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        for c in row:
            if c is None:
                continue
            txt = limpiar_texto(c).upper()
            txt = txt.replace("SECCION", "SECCIÓN")

            if txt.startswith("SECCIÓN "):
                m = re.search(r"SECCIÓN\s+([A-Z0-9\.]+)", txt)
                if m:
                    id_sec = m.group(1).upper()
                    id_sec = normalizar_seccion(id_sec)
                else:
                    id_sec = "?"

                secciones.append({
                    "fila_titulo": idx,
                    "id_seccion": id_sec,
                })
                break

    secciones.sort(key=lambda s: s["fila_titulo"])
    return secciones


def extraer_tabla_de_seccion(ws, fila_titulo, fila_fin_seccion):
    fila_header1_idx = None
    header_row1 = None
    header_row2 = None
    fila_header2_idx = None

    # buscar primera fila de encabezado
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if idx <= fila_titulo:
            continue
        if idx > fila_fin_seccion:
            break
        if es_fila_vacia(row):
            continue
        if es_fila_header(row):
            fila_header1_idx = idx
            header_row1 = [limpiar_texto(c) for c in row]
            break

    if header_row1 is None:
        return None, None, None, []

    # buscar segunda fila de encabezado
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if idx <= fila_header1_idx:
            continue
        if idx > fila_fin_seccion:
            break
        if es_fila_vacia(row):
            continue
        if es_fila_header(row):
            fila_header2_idx = idx
            header_row2 = [limpiar_texto(c) for c in row]
            break

    if header_row2 is None:
        header_row2 = [""] * len(header_row1)

    # filas de datos
    fila_inicio_datos = (fila_header2_idx or fila_header1_idx) + 1
    filas_datos = []
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if idx < fila_inicio_datos:
            continue
        if idx > fila_fin_seccion:
            break
        if es_fila_vacia(row):
            continue
        filas_datos.append(list(row))

    return fila_header1_idx, header_row1, header_row2, filas_datos


# ==========================
# FUNCIÓN PRINCIPAL
# ==========================

def procesar_archivo_con_mapeo(ruta_excel):
    mapeo = get_mapeo()
    wb = load_workbook(ruta_excel, data_only=True)
    registros = []

    for ws in wb.worksheets:
        titulo = ws.title.strip().upper()

        m = re.search(r"A[0-9]+[A-Z]?", titulo)
        if not m:
            continue
        hoja_codigo = m.group(0).upper()

        secciones = extraer_secciones_de_hoja(ws)
        if not secciones:
            continue

        for i, sec in enumerate(secciones):
            fila_titulo = sec["fila_titulo"]
            id_seccion = normalizar_seccion(sec["id_seccion"])

            if i + 1 < len(secciones):
                fila_fin = secciones[i + 1]["fila_titulo"] - 1
            else:
                fila_fin = ws.max_row

            fila_h1, header1, header2, filas_datos = extraer_tabla_de_seccion(
                ws, fila_titulo, fila_fin
            )
            if not filas_datos:
                continue

            max_len = max(len(header1), len(header2))

            columnas_utiles = []
            for col_idx in range(1, max_len + 1):
                col_letter = get_column_letter(col_idx).upper()
                clave = (hoja_codigo, id_seccion, col_letter)
                cfg = mapeo.get(clave)
                if not cfg:
                    continue

                campo_destino = cfg["campo_destino"].strip()
                if not campo_destino or campo_destino == "0":
                    continue

                columnas_utiles.append((col_idx, cfg))

            if not columnas_utiles:
                continue

            for idx_local, fila in enumerate(filas_datos, start=1):
                reg = {
                    "hoja": hoja_codigo,
                    "seccion": id_seccion,
                    "fila": idx_local,
                }
                for col_idx, cfg in columnas_utiles:
                    idx0 = col_idx - 1
                    valor = fila[idx0] if idx0 < len(fila) else None
                    tipo = cfg["tipo_dato"]
                    campo = cfg["campo_destino"].strip()

                    if valor is None:
                        reg[campo] = None
                    else:
                        if tipo == "entero":
                            try:
                                reg[campo] = int(valor)
                            except Exception:
                                try:
                                    reg[campo] = int(float(valor))
                                except Exception:
                                    reg[campo] = None
                        else:
                            reg[campo] = str(valor).strip()

                registros.append(reg)

    return registros
